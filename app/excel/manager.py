import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
import threading
from datetime import datetime
import logging
from dotenv import load_dotenv

load_dotenv()
logger = logging.getLogger(__name__)

# Thread lock — multiple bot users could write simultaneously
_excel_lock = threading.Lock()

# Resolve EXCEL_PATH to an absolute path at module load time.
# CRITICAL: If the project is inside OneDrive, OneDrive aggressively locks
# .xlsx files after every write (for cloud sync), causing PermissionError
# on the next write. To avoid this, we store the Excel file in a local
# directory OUTSIDE of OneDrive.
_BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_env_excel = os.getenv("EXCEL_FILE_PATH", "excel/appointments.xlsx")

# Detect if project is inside OneDrive
if "OneDrive" in _BASE_DIR:
    # Store Excel outside OneDrive to avoid file-locking issues
    _SAFE_DIR = os.path.join(os.path.expanduser("~"), "voice_bot_data")
    EXCEL_PATH = os.path.join(_SAFE_DIR, os.path.basename(_env_excel))
    logger.info(
        f"Project is inside OneDrive — Excel file redirected to: {EXCEL_PATH}"
    )
else:
    EXCEL_PATH = os.path.join(_BASE_DIR, _env_excel)

print(f"\n📁 Excel file location: {EXCEL_PATH}\n")

COLUMNS = [
    "appointment_id",
    "security_token",
    "patient_name",
    "patient_age",
    "patient_phone",
    "telegram_chat_id",
    "preferred_date",
    "preferred_time",
    "symptoms",
    "status",
    "confirmed_date",
    "confirmed_time",
    "doctor_notes",
    "created_at",
    "updated_at",
]

HEADER_FILL = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
HEADER_FONT = Font(bold=True)


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _ensure_workbook_exists():
    """Creates the Excel file with headers ONLY if it does not already exist."""
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    if not os.path.exists(EXCEL_PATH):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Appointments"

        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = max(
                15, len(col_name) + 4
            )

        wb.save(EXCEL_PATH)
        logger.info(f"Created new Excel file at: {EXCEL_PATH}")
    # If file already exists — do nothing. All appointments accumulate in this one file.


# ---------------------------------------------------------------------------
# CRUD operations
# ---------------------------------------------------------------------------

def write_appointment(appointment: dict) -> None:
    """
    Appends a new appointment row to the EXISTING Excel sheet.
    Thread-safe via _excel_lock.
    Retries up to 5 times (2s apart) in case the file is open in Excel.
    NEVER creates a new Workbook — only appends to the existing file.
    """
    import time

    _ensure_workbook_exists()

    max_retries = 5
    for attempt in range(max_retries):
        with _excel_lock:
            try:
                wb = openpyxl.load_workbook(EXCEL_PATH)
                ws = wb.active

                row_values = [appointment.get(col, "") for col in COLUMNS]
                ws.append(row_values)

                wb.save(EXCEL_PATH)
                logger.info(f"Appointment appended: {appointment.get('appointment_id')}")
                return  # Success

            except PermissionError as e:
                logger.warning(
                    f"Excel file locked (attempt {attempt+1}/{max_retries}): {e}"
                )
                if attempt < max_retries - 1:
                    time.sleep(2)
                else:
                    logger.error("All retries exhausted — file still locked.", exc_info=True)
                    raise

            except Exception as e:
                logger.error(f"Failed to write appointment: {e}", exc_info=True)
                raise


def get_appointment_by_id(appointment_id: str) -> dict | None:
    """
    Searches the Excel sheet for a row with the given appointment_id.
    Returns a dict of the row, or None if not found.
    """
    _ensure_workbook_exists()

    with _excel_lock:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == appointment_id:
                return dict(zip(COLUMNS, row))

    return None


def update_appointment_status(
    appointment_id: str,
    security_token: str,
    new_status: str,
    confirmed_date: str = "",
    confirmed_time: str = "",
    doctor_notes: str = "",
) -> dict | None:
    """
    Updates the status of an existing appointment row.
    Validates the security_token before making any changes.
    """
    _ensure_workbook_exists()

    with _excel_lock:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=2, values_only=False), start=2
        ):
            cell_id = row[0].value
            cell_token = row[1].value

            if cell_id == appointment_id:
                if cell_token != security_token:
                    logger.warning(
                        f"Token mismatch for appointment {appointment_id}. "
                        f"Expected: {cell_token}, Got: {security_token}"
                    )
                    return None

                now = datetime.now().isoformat(timespec="seconds")

                ws.cell(
                    row=row_idx, column=COLUMNS.index("status") + 1
                ).value = new_status
                ws.cell(
                    row=row_idx, column=COLUMNS.index("confirmed_date") + 1
                ).value = confirmed_date
                ws.cell(
                    row=row_idx, column=COLUMNS.index("confirmed_time") + 1
                ).value = confirmed_time
                ws.cell(
                    row=row_idx, column=COLUMNS.index("doctor_notes") + 1
                ).value = doctor_notes
                ws.cell(
                    row=row_idx, column=COLUMNS.index("updated_at") + 1
                ).value = now

                wb.save(EXCEL_PATH)
                logger.info(f"Appointment {appointment_id} updated to: {new_status}")

                updated_row = [cell.value for cell in ws[row_idx]]
                return dict(zip(COLUMNS, updated_row))

        logger.warning(f"Appointment ID not found: {appointment_id}")
        return None


# ---------------------------------------------------------------------------
# Conflict detection
# ---------------------------------------------------------------------------

def is_slot_taken(preferred_date: str, preferred_time: str) -> bool:
    """
    Returns True if there is already a Pending or Confirmed appointment
    at the given date and time. Rescheduled appointments are excluded.
    Retries up to 5 times if the file is locked.
    """
    import time

    _ensure_workbook_exists()

    max_retries = 5
    for attempt in range(max_retries):
        with _excel_lock:
            try:
                wb = openpyxl.load_workbook(EXCEL_PATH)
                ws = wb.active

                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = dict(zip(COLUMNS, row))
                    if (
                        row_dict.get("preferred_date") == preferred_date
                        and row_dict.get("preferred_time") == preferred_time
                        and row_dict.get("status") in ("Pending", "Confirmed")
                    ):
                        return True

                return False

            except PermissionError:
                logger.warning(
                    f"Excel locked during slot check (attempt {attempt+1}/{max_retries})"
                )
                if attempt < max_retries - 1:
                    time.sleep(2)
                else:
                    logger.error("Cannot check slot — file locked after all retries")
                    return False  # Assume free rather than crash

    return False
